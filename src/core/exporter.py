from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.core.models import VendedorResumen, ClienteResumen, ItemConsolidado

logger = logging.getLogger(__name__)

ESTILOS = {
    "header": {
        "font": {"bold": True, "color": "FFFFFF", "size": 11},
        "fill": {"patternType": "solid", "fgColor": "059669"},
        "alignment": {"horizontal": "center", "vertical": "center", "wrapText": True},
        "border": {"top": {"style": "thin"}, "bottom": {"style": "thin"},
                   "left": {"style": "thin"}, "right": {"style": "thin"}},
    },
    "total": {
        "font": {"bold": True, "size": 11, "color": "059669"},
        "fill": {"patternType": "solid", "fgColor": "ECFDF5"},
        "border": {"top": {"style": "thin"}, "bottom": {"style": "double"},
                   "left": {"style": "thin"}, "right": {"style": "thin"}},
    },
    "data": {
        "border": {"top": {"style": "thin"}, "bottom": {"style": "thin"},
                   "left": {"style": "thin"}, "right": {"style": "thin"}},
        "alignment": {"vertical": "center"},
    },
    "pct": {
        "border": {"top": {"style": "thin"}, "bottom": {"style": "thin"},
                   "left": {"style": "thin"}, "right": {"style": "thin"}},
        "alignment": {"horizontal": "center", "vertical": "center"},
        "number_format": "0.00%",
    },
    "soles": {
        "border": {"top": {"style": "thin"}, "bottom": {"style": "thin"},
                   "left": {"style": "thin"}, "right": {"style": "thin"}},
        "number_format": '#,##0.00',
    },
    "cantidad": {
        "border": {"top": {"style": "thin"}, "bottom": {"style": "thin"},
                   "left": {"style": "thin"}, "right": {"style": "thin"}},
        "number_format": '#,##0',
    },
}

COLUMNAS = ["SKU", "NOM_ARTICULO", "LINEA", "CANTIDAD", "SOLES", "% CANT", "% SOLES"]


def _generar_hoja_cliente(writer, df_hoja: pd.DataFrame, nom_cliente: str, nom_vendedor: str, used_names: set | None = None):
    sheet_name = _unique_sheet_name(nom_cliente[:31], used_names or set())
    df_hoja = df_hoja.groupby("SKU", as_index=False).agg({
        "NOM_ARTICULO": "first", "LINEA": "first",
        "CANTIDAD": "sum", "SOLES": "sum",
    })

    total_soles = df_hoja["SOLES"].sum()

    if total_soles > 0:
        df_hoja["% SOLES"] = df_hoja["SOLES"] / total_soles
    else:
        df_hoja["% SOLES"] = 0.0

    df_hoja = df_hoja.sort_values(["LINEA", "SKU"], ascending=[True, True])
    df_hoja = df_hoja.reset_index(drop=True)

    n_data = len(df_hoja)
    total_row = n_data + 2

    BASE_COLUMNAS = ["SKU", "NOM_ARTICULO", "LINEA", "CANTIDAD", "SOLES"]
    df_out = df_hoja[BASE_COLUMNAS].copy()
    df_out.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    ws = writer.sheets[sheet_name]

    ws.cell(row=1, column=6, value="% CANT")
    ws.cell(row=1, column=7, value="% SOLES")

    for i in range(n_data):
        r = i + 2
        ws.cell(row=r, column=6).value = f"=IFERROR(D{r}/D${total_row}, 0)"
        ws.cell(row=r, column=7).value = f"=IFERROR(E{r}/E${total_row}, 0)"

    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value="")
    ws.cell(row=total_row, column=3, value="")
    ws.cell(row=total_row, column=4).value = f"=SUM(D2:D{n_data+1})"
    ws.cell(row=total_row, column=5).value = f"=SUM(E2:E{n_data+1})"
    ws.cell(row=total_row, column=6).value = 1.0
    ws.cell(row=total_row, column=7).value = 1.0

    table_ref = f"A1:G{n_data+1}"
    table_name = re.sub(r'[^a-zA-Z0-9_]', '', f"tbl_{sheet_name[:20]}")
    tab = Table(displayName=table_name, ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Side(style="thin")
    double_bottom = Side(style="double")

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="059669")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    total_font = Font(bold=True, size=11, color="059669")
    total_fill = PatternFill("solid", fgColor="ECFDF5")
    total_border = Border(top=thin, bottom=double_bottom, left=thin, right=thin)

    data_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    pct_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(COLUMNAS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = hdr_border

    for row_idx in range(2, n_data + 2):
        for col_idx, col_name in enumerate(COLUMNAS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            if col_name in ("% CANT", "% SOLES"):
                cell.alignment = pct_align
                cell.number_format = "0.00%"
            elif col_name == "SOLES":
                cell.number_format = "#,##0.00"
            elif col_name == "CANTIDAD":
                cell.number_format = "#,##0"

    for col_idx, col_name in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        if col_name in ("% CANT", "% SOLES"):
            cell.number_format = "0.00%"
        elif col_name == "SOLES":
            cell.number_format = "#,##0.00"
        elif col_name == "CANTIDAD":
            cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    widths = [15, 45, 20, 12, 14, 10, 10]
    for i, w in enumerate(widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

def _unique_sheet_name(base: str, used: set) -> str:
    name = base
    i = 1
    while name in used:
        suffix = f" ({i})"
        max_base = 31 - len(suffix)
        name = base[:max_base] + suffix
        i += 1
    used.add(name)
    return name


def exportar_xlsx(por_vendedor: dict[str, VendedorResumen],
                  clientes_seleccionados: dict[str, list[str]],
                  lineas_sel: set[str] | None = None,
                  carpeta: Optional[str] = None,
                  incluir_sucursales: bool = False,
                  filter_items: Optional[callable] = None) -> list[str]:
    if carpeta is None:
        carpeta = str(Path.home() / "Desktop")

    generados = []
    fecha = datetime.now().strftime("%d%m%Y")

    for vid, vendedor in por_vendedor.items():
        clientes = clientes_seleccionados.get(vid, [])
        if not clientes:
            continue

        nom_ven = re.sub(r'[<>:"/\\|?*]', '', vendedor.nom_vendedor).replace(" ", "_")[:20]
        fname = f"G360_Consolidado_{nom_ven}_{fecha}.xlsx"
        fpath = os.path.join(carpeta, fname)

        with pd.ExcelWriter(fpath, engine="openpyxl") as writer:
            writer.book.properties.creator = "ccusi"
            writer.book.properties.title = "PreOrder Allocator"
            writer.book.properties.description = "Generado por G360"
            used_sheets: set[str] = set()
            for nom_cliente in sorted(clientes):
                cliente = vendedor.clientes.get(nom_cliente)
                if not cliente or not cliente.items:
                    continue

                items = cliente.items
                if lineas_sel:
                    items = [it for it in items if f"{it.id_linea} - {it.nom_linea}" in lineas_sel]
                if filter_items:
                    items = filter_items(items)
                if not items:
                    continue

                data = []
                for item in items:
                    data.append({
                        "SKU": item.sku,
                        "NOM_ARTICULO": item.nom_articulo,
                        "LINEA": item.nom_linea,
                        "CANTIDAD": item.cantidad,
                        "SOLES": item.soles,
                    })

                df_hoja = pd.DataFrame(data)
                _generar_hoja_cliente(writer, df_hoja, nom_cliente, vendedor.nom_vendedor, used_names=used_sheets)

                if incluir_sucursales:
                    _generar_hoja_sucursales(writer, items, nom_cliente, used_names=used_sheets)

        generados.append(fpath)
        logger.info(f"Reporte generado: {fpath}")

    return generados


def _generar_hoja_sucursales(writer, items: list[ItemConsolidado], nom_cliente: str, used_names: set | None = None):
    suc_data = {}
    for it in items:
        cod = it.cod_sucursal or "S/N"
        nom = it.nom_sucursal or "Sin Sucursal"
        key = f"{cod} - {nom}"
        if key not in suc_data:
            suc_data[key] = {"cantidad": 0.0, "soles": 0.0}
        suc_data[key]["cantidad"] += it.cantidad
        suc_data[key]["soles"] += it.soles

    if not suc_data:
        return

    rows = []
    total_cant = sum(v["cantidad"] for v in suc_data.values())
    total_soles = sum(v["soles"] for v in suc_data.values())

    for key in sorted(suc_data.keys()):
        v = suc_data[key]
        rows.append({
            "SUCURSAL": key,
            "CANTIDAD": int(v["cantidad"]),
            "SOLES": round(v["soles"], 2),
            "% CANT": v["cantidad"] / total_cant if total_cant > 0 else 0,
            "% SOLES": v["soles"] / total_soles if total_soles > 0 else 0,
        })

    if total_cant > 0:
        rows.append({
            "SUCURSAL": "TOTAL",
            "CANTIDAD": int(total_cant),
            "SOLES": round(total_soles, 2),
            "% CANT": 1.0,
            "% SOLES": 1.0,
        })

    sheet_name = _unique_sheet_name(f"Suc. {nom_cliente}"[:31], used_names or set())
    df_suc = pd.DataFrame(rows)
    df_suc.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    ws = writer.sheets[sheet_name]

    ncols = len(df_suc.columns)
    ncols_letter = get_column_letter(ncols)
    table_ref = f"A1:{ncols_letter}{len(df_suc)}"
    table_name = re.sub(r'[^a-zA-Z0-9_]', '', f"tbl_{sheet_name[:20]}")
    tab = Table(displayName=table_name, ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Side(style="thin")
    double_bottom = Side(style="double")

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="059669")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    total_font = Font(bold=True, size=11, color="059669")
    total_fill = PatternFill("solid", fgColor="ECFDF5")
    total_border = Border(top=thin, bottom=double_bottom, left=thin, right=thin)

    data_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    pct_align = Alignment(horizontal="center", vertical="center")

    col_count = len(df_suc.columns)
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = hdr_border

    for row_idx in range(2, len(df_suc) + 2):
        is_total = row_idx == len(df_suc) + 1
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_total:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = total_border
            else:
                cell.border = data_border
                col_name = df_suc.columns[col_idx - 1]
                if col_name in ("% CANT", "% SOLES"):
                    cell.alignment = pct_align
                    cell.number_format = "0.00%"
                elif col_name == "SOLES":
                    cell.number_format = "#,##0.00"
                elif col_name == "CANTIDAD":
                    cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    widths = [30, 12, 14, 10, 10]
    for i, w in enumerate(widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

